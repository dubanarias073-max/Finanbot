# routes/excel.py 
# Fix principal: categoria es string directo, NO objeto con .nombre
from flask import Blueprint, send_file, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import Usuario, Transaccion, MetaAhorro, Simulacion
from datetime import datetime
from collections import defaultdict
import io

# ── IMPORTS DE OPENPYXL OPTIMIZADOS AL INICIO ─────────────────
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter

excel_bp = Blueprint('excel', __name__)
 
# ── Iconos por categoría (string directo) ─────────────────
ICONOS = {
    'Alimentación':'🍔','Transporte':'🚌','Arriendo':'🏠','Salud':'💊',
    'Entretenimiento':'🎬','Educación':'📚','Ropa':'👗','Servicios':'⚡',
    'Tecnología':'💻','Médico':'🏥','Mascotas':'🐾','Regalos':'🎁',
    'Restaurantes':'🍕','Viajes':'✈️','Otros gastos':'📦','Salario':'💼',
    'Freelance':'🧑‍💻','Inversión':'📈','Negocio':'🏪','Regalo':'🎁',
    'Otros ingresos':'💵',
}
 
def get_cat(t):
    """Devuelve la categoría siempre como string legible."""
    cat = t.categoria
    if cat is None:
        return 'Sin categoría'
    if isinstance(cat, str):
        return cat
    # Por si en algún momento es objeto SQLAlchemy
    if hasattr(cat, 'nombre'):
        return cat.nombre
    return str(cat)
 
def get_icono(cat_str):
    return ICONOS.get(cat_str, '💸')
 
def tipo_inv(tasa):
    t = float(tasa)
    if t <= 5:  return '🏦 CDT Básico'
    if t <= 8:  return '💎 CDT Premium'
    if t <= 10: return '📊 Fondo Inversión'
    if t <= 15: return '📈 Acciones BVC'
    if t <= 20: return '🚀 Startups'
    return f'⚙️ Personalizada ({t}%)'
 
 
@excel_bp.route('/excel', methods=['GET'])
@jwt_required()
def exportar_excel():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.chart import BarChart, PieChart, LineChart, Reference
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'mensaje': 'Instala openpyxl: pip install openpyxl'}), 500
 
    uid  = int(get_jwt_identity())
    user = Usuario.query.get(uid)
    if not user:
        return jsonify({'mensaje': 'Usuario no encontrado'}), 404
 
    trans = Transaccion.query.filter_by(usuario_id=uid).order_by(Transaccion.fecha.asc()).all()
    metas = MetaAhorro.query.filter_by(usuario_id=uid).all()
    sims  = Simulacion.query.filter_by(usuario_id=uid).order_by(Simulacion.fecha.asc()).all()
 
    # ── Paleta ────────────────────────────────────────────
    P = {
        'mo':'C026D3','md':'7E22CE','cy':'06B6D4','ve':'22C55E',
        'ro':'F472B6','am':'F59E0B','re':'EF4444','bg':'0F0F2D',
        'ca':'1A1A4E','li':'2D1B69','bl':'FFFFFF','gr':'9CA3AF',
        'mu':'6B7280','az':'0EA5E9','os':'1E293B',
    }
 
    wb = openpyxl.Workbook()
 
# ── Helpers ───────────────────────────────────────────
    def fl(h):  return PatternFill('solid', fgColor=h)
    def fn(h=None, bold=False, sz=10, italic=False):
        return Font(color=h or P['bl'], bold=bold, size=sz,
                    italic=italic, name='Segoe UI')
    # CORRECCIÓN 1: Cambiado 'wrap_text' por 'wrapText' (así lo exige OpenPyXL)
    def al(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrapText=wrap)
    def bd(col=None):
        s = Side(style='thin', color=col or P['li'])
        return Border(left=s, right=s, top=s, bottom=s)
    def cw(ws, col, w):
        ws.column_dimensions[get_column_letter(col)].width = w
    def rh(ws, row, h):
        ws.row_dimensions[row].height = h
 
    def wr(ws, row, col, val, bg=None, fg=None, bold=False,
           sz=10, h='center', brd=True, italic=False, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fl(bg or P['ca'])
        c.font = fn(fg or P['bl'], bold, sz, italic)
        # CORRECCIÓN 2: Cambiado 'wrap_text=wrap' por 'wrap=wrap' (para coincidir con la firma de tu función al)
        c.alignment = al(h, wrap=wrap)
        if brd: c.border = bd()
        return c
 
    def fondo_hoja(ws, filas=300, cols=20):
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows(min_row=1, max_row=filas, min_col=1, max_col=cols):
            for c in row: c.fill = fl(P['bg'])
 
    def cabecera(ws, titulo, sub, cols=8):
        fondo_hoja(ws)
        rh(ws, 1, 46); rh(ws, 2, 20)
        ws.merge_cells(f'A1:{get_column_letter(cols)}1')
        c1 = ws['A1']; c1.value = titulo
        c1.fill = fl(P['mo']); c1.font = fn(P['bl'], bold=True, sz=18)
        c1.alignment = al()
        ws.merge_cells(f'A2:{get_column_letter(cols)}2')
        c2 = ws['A2']; c2.value = sub
        c2.fill = fl(P['md']); c2.font = fn(P['gr'], sz=9, italic=True)
        c2.alignment = al()
 
    def enc_fila(ws, row, defs, bg=None):
        rh(ws, row, 26)
        for i, (txt, ancho) in enumerate(defs, 1):
            c = ws.cell(row=row, column=i, value=txt)
            c.fill = fl(bg or P['li'])
            c.font = fn(P['bl'], bold=True, sz=9)
            c.alignment = al(); c.border = bd(); cw(ws, i, ancho)
 
    def kpi_block(ws, row_start, items):
        """Dibuja tarjetas KPI: label arriba, valor grande abajo."""
        rh(ws, row_start,   18)
        rh(ws, row_start+1, 36)
        rh(ws, row_start+2, 12)
        for i, (lbl, val, col) in enumerate(items, 1):
            for r_off, tx, sz_, bld in [
                (0, lbl, 8, False), (1, val, 14, True), (2, '', 7, False)
            ]:
                c = ws.cell(row=row_start+r_off, column=i, value=tx)
                c.fill = fl(P['bg'] if r_off == 0 else P['ca'])
                c.font = fn(P['mu'] if r_off == 0 else col, bld, sz_)
                c.alignment = al(); c.border = bd(col)
            cw(ws, i, 18)

 
    # ── Métricas globales ─────────────────────────────────
    ing_t = sum(float(t.monto) for t in trans if t.tipo == 'ingreso')
    gas_t = sum(float(t.monto) for t in trans if t.tipo == 'gasto')
    bal_t = ing_t - gas_t
    aho_t = sum(float(m.monto_actual) for m in metas)
    pct_g = round(gas_t / ing_t * 100) if ing_t > 0 else 0
 
    # ← FIX: get_cat() garantiza string legible
    gastos_cat   = defaultdict(float)
    ingresos_cat = defaultdict(float)
    cat_stats    = defaultdict(lambda: {'total':0,'count':0,'max':0})
    ing_stats    = defaultdict(lambda: {'total':0,'count':0,'max':0})
 
    for t in trans:
        nombre_cat = get_cat(t)          # ← SIEMPRE string
        mn = float(t.monto)
        if t.tipo == 'gasto':
            gastos_cat[nombre_cat]          += mn
            cat_stats[nombre_cat]['total']  += mn
            cat_stats[nombre_cat]['count']  += 1
            cat_stats[nombre_cat]['max']     = max(cat_stats[nombre_cat]['max'], mn)
        else:
            ingresos_cat[nombre_cat]        += mn
            ing_stats[nombre_cat]['total']  += mn
            ing_stats[nombre_cat]['count']  += 1
            ing_stats[nombre_cat]['max']     = max(ing_stats[nombre_cat]['max'], mn)
 
    ing_mens   = float(user.ingreso_mensual or 0)
    meta_mens  = float(user.meta_ahorro or 0)
    metas_comp = [m for m in metas if m.completada]
    cat_mayor  = max(gastos_cat, key=gastos_cat.get) if gastos_cat else None
 
    # ════════════════════════════════════════════════════
    #  HOJA 1 — RESUMEN EJECUTIVO
    # ════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = 'Resumen'
    ws1.sheet_properties.tabColor = P['mo']
    cabecera(ws1,
        'FinanBot — Reporte Financiero',
        f'Usuario: {user.nombre}  ·  {user.correo}  ·  {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        8)
 
    kpi_block(ws1, 4, [
        ('INGRESOS',     f'${ing_t:,.0f}',  P['cy']),
        ('GASTOS',       f'${gas_t:,.0f}',  P['mo']),
        ('BALANCE',      f'${bal_t:,.0f}',  P['ve'] if bal_t >= 0 else P['re']),
        ('AHORRADO',     f'${aho_t:,.0f}',  P['am']),
        ('TRANSAC.',     str(len(trans)),     P['gr']),
        ('% GASTOS',     f'{pct_g}%',        P['ve'] if pct_g < 70 else P['am'] if pct_g < 90 else P['re']),
        ('METAS',        str(len(metas)),     P['cy']),
        ('SIMULAC.',     str(len(sims)),      P['gr']),
    ])
 
    # Separador
    rh(ws1, 7, 5)
    for i in range(1, 9):
        ws1.cell(row=7, column=i).fill = fl(P['li'])
 
    # Indicadores
    enc_fila(ws1, 8, [('Indicador', 30), ('Valor', 16), ('Estado', 28)])
 
    inds = [
        ('% Gastos sobre ingresos', f'{pct_g}%',
         '✅ Bueno' if pct_g < 70 else '⚠️ Ajustado' if pct_g < 90 else '🚨 Crítico'),
        (f'Metas completadas', f'{len(metas_comp)} de {len(metas)}',
         '✅ Excelente' if metas_comp else '—'),
    ]
    if cat_mayor:
        pm = round(gastos_cat[cat_mayor] / gas_t * 100) if gas_t else 0
        inds.append((f'Mayor gasto: {get_icono(cat_mayor)} {cat_mayor}',
                     f'${gastos_cat[cat_mayor]:,.0f} ({pm}%)',
                     '⚠️ Alta' if pm > 40 else '✅ Normal'))
    if ing_mens > 0:
        inds += [
            ('50% Necesidades (regla 50/30/20)', f'${ing_mens*.5:,.0f}', 'Máx. recomendado/mes'),
            ('20% Ahorro/Inversión',             f'${ing_mens*.2:,.0f}', 'Meta mensual ideal'),
        ]
    if meta_mens > 0:
        inds.append(('Meta de ahorro mensual', f'${meta_mens:,.0f}',
                     '✅ Cumplida' if bal_t >= meta_mens else f'⚠️ Brecha ${max(0, meta_mens-bal_t):,.0f}'))
    fondo_obj = ing_mens * 3 if ing_mens > 0 else ing_t * 3
    pct_f = min(100, round(max(0, bal_t) / fondo_obj * 100)) if fondo_obj > 0 else 0
    inds.append(('Fondo de emergencia (3 meses)',
                 f'{pct_f}%', '✅ Completo' if pct_f >= 100 else f'Meta: ${fondo_obj:,.0f}'))
 
    for ri, (ind, val, est) in enumerate(inds, 9):
        rh(ws1, ri, 20)
        bg_ = P['ca'] if ri % 2 == 0 else P['bg']
        ec  = P['ve'] if '✅' in est else P['am'] if '⚠️' in est else P['re'] if '🚨' in est else P['gr']
        wr(ws1, ri, 1, ind, bg_, P['bl'], sz=9, h='left')
        wr(ws1, ri, 2, val, bg_, P['cy'], sz=9)
        wr(ws1, ri, 3, est, bg_, ec,      sz=9)
 
    # Gráfico gastos
    ws1['J1'] = 'Categoría'; ws1['K1'] = 'Monto'
    ws1['J1'].fill = fl(P['li']); ws1['J1'].font = fn(P['bl'], bold=True, sz=9)
    ws1['K1'].fill = fl(P['li']); ws1['K1'].font = fn(P['bl'], bold=True, sz=9)
    cw(ws1, 10, 26); cw(ws1, 11, 14)
    rg = 2
    for cat, monto in sorted(gastos_cat.items(), key=lambda x: x[1], reverse=True)[:10]:
        ws1[f'J{rg}'] = f'{get_icono(cat)} {cat}'   # ← nombre real de categoría
        ws1[f'K{rg}'] = round(monto, 2)
        ws1[f'J{rg}'].fill = fl(P['ca']); ws1[f'J{rg}'].font = fn(P['bl'], sz=9)
        ws1[f'K{rg}'].fill = fl(P['ca']); ws1[f'K{rg}'].font = fn(P['am'], sz=9)
        rg += 1
    if rg > 2:
        bar = BarChart(); bar.type = 'col'; bar.title = 'Gastos por Categoría'
        bar.style = 10; bar.width = 20; bar.height = 13
        dr = Reference(ws1, min_col=11, min_row=1, max_row=rg-1)
        cr = Reference(ws1, min_col=10, min_row=2, max_row=rg-1)
        bar.add_data(dr, titles_from_data=True); bar.set_categories(cr)
        bar.series[0].graphicalProperties.solidFill = P['mo']
        ws1.add_chart(bar, 'J4')
 
    # ════════════════════════════════════════════════════
    #  HOJA 2 — TRANSACCIONES
    # ════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Transacciones')
    ws2.sheet_properties.tabColor = P['cy']
    cabecera(ws2, 'Historial de Transacciones',
        f'{len(trans)} registros  ·  Ingresos: ${ing_t:,.0f}  ·  Gastos: ${gas_t:,.0f}  ·  Balance: ${bal_t:,.0f}',
        9)
    enc_fila(ws2, 4, [
        ('#', 5), ('Fecha', 13), ('Tipo', 11),
        ('Categoría', 22),            # ← nombre real
        ('Descripción', 28),
        ('Monto', 16), ('Saldo acum.', 16), ('Retención', 12), ('Neto', 14),
    ])
 
    saldo = 0.0
    for idx, t in enumerate(trans, 1):
        nombre_cat = get_cat(t)      # ← FIX: siempre string
        mn   = float(t.monto)
        saldo += mn if t.tipo == 'ingreso' else -mn
        r    = idx + 4
        bg_  = P['ca'] if idx % 2 == 0 else P['bg']
        col_ = P['ve'] if t.tipo == 'ingreso' else P['ro']
        sgn  = '+' if t.tipo == 'ingreso' else '-'
        fd   = t.fecha.strftime('%d/%m/%Y') if t.fecha else '—'
        ret_ = round(mn * 0.04, 2) if t.tipo == 'ingreso' else 0
        rh(ws2, r, 19)
        wr(ws2, r, 1, idx,                                  bg_, P['mu'], sz=9)
        wr(ws2, r, 2, fd,                                   bg_, P['gr'], sz=9)
        wr(ws2, r, 3, t.tipo.capitalize(),                  bg_, col_, bold=True, sz=9)
        wr(ws2, r, 4, f'{get_icono(nombre_cat)} {nombre_cat}', bg_, P['bl'], sz=9, h='left')
        wr(ws2, r, 5, t.descripcion or '—',                 bg_, P['gr'], sz=9, h='left')
        wr(ws2, r, 6, f'{sgn}${mn:,.0f}',                  bg_, col_, bold=True, sz=10)
        wr(ws2, r, 7, f'${saldo:,.0f}',                    bg_, P['cy'] if saldo >= 0 else P['re'], sz=9)
        wr(ws2, r, 8, f'${ret_:,.0f}' if ret_ > 0 else '—', bg_, P['ro'], sz=9)
        wr(ws2, r, 9, f'${mn - ret_:,.0f}' if t.tipo == 'ingreso' else f'-${mn:,.0f}',
           bg_, P['ve'] if t.tipo == 'ingreso' else P['ro'], sz=9)
 
    ws2.auto_filter.ref = 'A4:I4'
 
    # Gráfico torta
    if gastos_cat:
        ws2['K1'] = 'Categoría'; ws2['L1'] = 'Monto'
        ws2['K1'].fill = fl(P['li']); ws2['K1'].font = fn(P['bl'], bold=True, sz=9)
        ws2['L1'].fill = fl(P['li']); ws2['L1'].font = fn(P['bl'], bold=True, sz=9)
        cw(ws2, 11, 24); cw(ws2, 12, 14)
        fp = 2
        for cat, monto in sorted(gastos_cat.items(), key=lambda x: x[1], reverse=True)[:8]:
            ws2[f'K{fp}'] = f'{get_icono(cat)} {cat}'   # ← nombre real
            ws2[f'L{fp}'] = round(monto, 2)
            ws2[f'K{fp}'].fill = fl(P['ca']); ws2[f'K{fp}'].font = fn(P['bl'], sz=9)
            ws2[f'L{fp}'].fill = fl(P['ca']); ws2[f'L{fp}'].font = fn(P['am'], sz=9)
            fp += 1
        pie = PieChart(); pie.title = 'Distribución de Gastos'
        pie.style = 10; pie.width = 18; pie.height = 13
        pd_ = Reference(ws2, min_col=12, min_row=1, max_row=fp-1)
        pc_ = Reference(ws2, min_col=11, min_row=2, max_row=fp-1)
        pie.add_data(pd_, titles_from_data=True); pie.set_categories(pc_)
        ws2.add_chart(pie, 'K4')
 
    # ════════════════════════════════════════════════════
    #  HOJA 3 — ANÁLISIS POR CATEGORÍA
    # ════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Análisis Categorías')
    ws3.sheet_properties.tabColor = P['am']
    cabecera(ws3, 'Análisis Detallado por Categoría',
        f'{len(gastos_cat)} categorías de gasto  ·  {len(ingresos_cat)} fuentes de ingreso', 7)
 
    # Gastos
    wr(ws3, 4, 1, '💸  GASTOS POR CATEGORÍA', P['li'], P['mo'], bold=True, sz=11, h='left')
    enc_fila(ws3, 5, [
        ('Categoría', 24), ('Total', 16), ('% del total', 12),
        ('# Transac.', 12), ('Promedio', 14), ('Máximo', 14), ('Nivel', 14),
    ], P['re'])
 
    for i, (cat, st) in enumerate(sorted(cat_stats.items(), key=lambda x: x[1]['total'], reverse=True), 6):
        bg_     = P['ca'] if i % 2 == 0 else P['bg']
        pct_cat = round(st['total'] / gas_t * 100, 1) if gas_t else 0
        prom    = round(st['total'] / st['count'], 0) if st['count'] else 0
        nivel   = '🚨 Alta' if pct_cat > 40 else '⚠️ Media' if pct_cat > 25 else '✅ Normal'
        nc      = P['re'] if pct_cat > 40 else P['am'] if pct_cat > 25 else P['ve']
        rh(ws3, i, 19)
        wr(ws3, i, 1, f'{get_icono(cat)} {cat}',  bg_, P['bl'], sz=9, h='left')  # ← nombre real
        wr(ws3, i, 2, f'${st["total"]:,.0f}',     bg_, P['ro'], bold=True, sz=9)
        wr(ws3, i, 3, f'{pct_cat}%',              bg_, P['am'], sz=9)
        wr(ws3, i, 4, st['count'],                 bg_, P['gr'], sz=9)
        wr(ws3, i, 5, f'${prom:,.0f}',            bg_, P['cy'], sz=9)
        wr(ws3, i, 6, f'${st["max"]:,.0f}',       bg_, P['gr'], sz=9)
        wr(ws3, i, 7, nivel,                       bg_, nc, bold=True, sz=9)
 
    # Ingresos
    fi = len(cat_stats) + 8
    wr(ws3, fi, 1, '💰  INGRESOS POR FUENTE', P['li'], P['cy'], bold=True, sz=11, h='left')
    enc_fila(ws3, fi+1, [
        ('Fuente', 24), ('Total', 16), ('% del total', 12),
        ('# Transac.', 12), ('Promedio', 14), ('Máximo', 14), ('Tendencia', 14),
    ], P['ve'])
 
    for i, (cat, st) in enumerate(sorted(ing_stats.items(), key=lambda x: x[1]['total'], reverse=True), fi+2):
        bg_     = P['ca'] if i % 2 == 0 else P['bg']
        pct_cat = round(st['total'] / ing_t * 100, 1) if ing_t else 0
        prom    = round(st['total'] / st['count'], 0) if st['count'] else 0
        tend    = '📈 Principal' if pct_cat > 50 else '📊 Secundario'
        rh(ws3, i, 19)
        wr(ws3, i, 1, f'{get_icono(cat)} {cat}',  bg_, P['bl'], sz=9, h='left')  # ← nombre real
        wr(ws3, i, 2, f'${st["total"]:,.0f}',     bg_, P['ve'], bold=True, sz=9)
        wr(ws3, i, 3, f'{pct_cat}%',              bg_, P['am'], sz=9)
        wr(ws3, i, 4, st['count'],                 bg_, P['gr'], sz=9)
        wr(ws3, i, 5, f'${prom:,.0f}',            bg_, P['cy'], sz=9)
        wr(ws3, i, 6, f'${st["max"]:,.0f}',       bg_, P['gr'], sz=9)
        wr(ws3, i, 7, tend,                        bg_, P['cy'], sz=9)
 
    # Gráfico barras horizontal gastos
    ws3['I1'] = 'Categoría'; ws3['J1'] = 'Monto'
    ws3['I1'].fill = fl(P['li']); ws3['I1'].font = fn(P['bl'], bold=True, sz=9)
    ws3['J1'].fill = fl(P['li']); ws3['J1'].font = fn(P['bl'], bold=True, sz=9)
    cw(ws3, 9, 26); cw(ws3, 10, 14)
    rc = 2
    for cat, st in sorted(cat_stats.items(), key=lambda x: x[1]['total'], reverse=True)[:8]:
        ws3[f'I{rc}'] = f'{get_icono(cat)} {cat}'   # ← nombre real
        ws3[f'J{rc}'] = round(st['total'], 2)
        ws3[f'I{rc}'].fill = fl(P['ca']); ws3[f'I{rc}'].font = fn(P['bl'], sz=9)
        ws3[f'J{rc}'].fill = fl(P['ca']); ws3[f'J{rc}'].font = fn(P['ro'], sz=9)
        rc += 1
    if rc > 2:
        bc = BarChart(); bc.type = 'bar'; bc.title = 'Ranking de Gastos'
        bc.style = 10; bc.width = 20; bc.height = 14
        dc = Reference(ws3, min_col=10, min_row=1, max_row=rc-1)
        cc = Reference(ws3, min_col=9,  min_row=2, max_row=rc-1)
        bc.add_data(dc, titles_from_data=True); bc.set_categories(cc)
        bc.series[0].graphicalProperties.solidFill = P['mo']
        ws3.add_chart(bc, 'I4')
 
    # ════════════════════════════════════════════════════
    #  HOJA 4 — METAS DE AHORRO
    # ════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Metas de Ahorro')
    ws4.sheet_properties.tabColor = P['ve']
    compl_ = len(metas_comp)
    cabecera(ws4, 'Metas de Ahorro',
        f'{len(metas)} metas  ·  {compl_} completadas  ·  Total ahorrado: ${aho_t:,.0f}', 8)
 
    kpi_block(ws4, 4, [
        ('TOTAL METAS',   str(len(metas)),   P['gr']),
        ('COMPLETADAS',   str(compl_),        P['ve']),
        ('EN PROGRESO',   str(len(metas)-compl_), P['am']),
        ('AHORRADO',      f'${aho_t:,.0f}',  P['cy']),
        ('OBJETIVO',      f'${sum(float(m.monto_objetivo) for m in metas):,.0f}', P['mo']),
    ])
 
    enc_fila(ws4, 8, [
        ('Meta', 28), ('Ahorrado', 15), ('Objetivo', 15),
        ('Faltante', 15), ('%', 9), ('Estado', 16),
        ('Prioridad', 13), ('Fecha límite', 14),
    ])
 
    for idx, m in enumerate(sorted(metas, key=lambda x: not x.completada), 9):
        bg_  = P['ca'] if idx % 2 == 0 else P['bg']
        pct_ = round(float(m.monto_actual)/float(m.monto_objetivo)*100, 1) if m.monto_objetivo else 0
        flt_ = max(0, float(m.monto_objetivo) - float(m.monto_actual))
        col_ = P['ve'] if m.completada else (P['cy'] if pct_ >= 75 else P['am'])
        est_ = '✅ Completada' if m.completada else ('🔥 Casi lista' if pct_ >= 75 else '⏳ En progreso')
        fl__ = m.fecha_limite.strftime('%d/%m/%Y') if m.fecha_limite else '—'
        prio = getattr(m, 'prioridad', '—') or '—'
        rh(ws4, idx, 21)
        wr(ws4, idx, 1, m.nombre,                         bg_, P['bl'], sz=9, h='left')
        wr(ws4, idx, 2, f'${float(m.monto_actual):,.0f}', bg_, P['ve'], bold=True, sz=9)
        wr(ws4, idx, 3, f'${float(m.monto_objetivo):,.0f}',bg_, P['cy'], sz=9)
        wr(ws4, idx, 4, f'${flt_:,.0f}',                  bg_, P['ro'], sz=9)
        wr(ws4, idx, 5, f'{pct_}%',                        bg_, P['am'], bold=True, sz=9)
        wr(ws4, idx, 6, est_,                              bg_, col_, bold=True, sz=9)
        wr(ws4, idx, 7, prio,                              bg_, P['gr'], sz=9)
        wr(ws4, idx, 8, fl__,                              bg_, P['gr'], sz=9)
 
    ws4.auto_filter.ref = 'A8:H8'
 
    if metas:
        ws4['J1'] = 'Meta'; ws4['K1'] = 'Ahorrado'; ws4['L1'] = 'Objetivo'
        for c in ['J1','K1','L1']:
            ws4[c].fill = fl(P['li']); ws4[c].font = fn(P['bl'], bold=True, sz=9)
        cw(ws4,10,24); cw(ws4,11,14); cw(ws4,12,14)
        rm = 2
        for m in metas[:10]:
            ws4[f'J{rm}'] = m.nombre[:22]
            ws4[f'K{rm}'] = round(float(m.monto_actual), 2)
            ws4[f'L{rm}'] = round(float(m.monto_objetivo), 2)
            for c2 in [f'J{rm}',f'K{rm}',f'L{rm}']:
                ws4[c2].fill = fl(P['ca']); ws4[c2].font = fn(P['bl'], sz=9)
            rm += 1
        bm = BarChart(); bm.type = 'col'; bm.title = 'Progreso de Metas'
        bm.style = 10; bm.width = 22; bm.height = 14
        dm = Reference(ws4, min_col=11, max_col=12, min_row=1, max_row=rm-1)
        cm = Reference(ws4, min_col=10, min_row=2, max_row=rm-1)
        bm.add_data(dm, titles_from_data=True); bm.set_categories(cm)
        bm.series[0].graphicalProperties.solidFill = P['ve']
        bm.series[1].graphicalProperties.solidFill = P['mo']
        ws4.add_chart(bm, 'J4')
 
    # ════════════════════════════════════════════════════
    #  HOJA 5 — SIMULACIONES
    # ════════════════════════════════════════════════════
    ws5 = wb.create_sheet('Simulaciones')
    ws5.sheet_properties.tabColor = P['az']
    cabecera(ws5, 'Simulaciones de Inversión',
        f'{len(sims)} simulaciones  ·  Interés compuesto mensual  ·  Retención 4% sobre ganancias', 9)
    enc_fila(ws5, 4, [
        ('#',5), ('Fecha',13), ('Tipo de inversión',20),
        ('Capital',16), ('Tasa %',12), ('Plazo',10),
        ('Resultado',16), ('Ganancia',16), ('Neto (c/ret.)',15),
    ])
 
    for idx, s in enumerate(sims, 1):
        # ← Campos correctos del modelo
        capital   = float(s.capital_inicial)
        resultado = float(s.resultado_final)
        tasa      = float(s.tasa_retorno)
        plazo     = int(s.plazo_meses)
        ganancia  = resultado - capital
        retencion = round(max(ganancia, 0) * 0.04, 2)
        neto      = resultado - retencion
        r   = idx + 4
        bg_ = P['ca'] if idx % 2 == 0 else P['bg']
        fd  = s.fecha.strftime('%d/%m/%Y') if s.fecha else '—'
        rh(ws5, r, 19)
        wr(ws5, r, 1, idx,                 bg_, P['mu'], sz=9)
        wr(ws5, r, 2, fd,                  bg_, P['gr'], sz=9)
        wr(ws5, r, 3, tipo_inv(tasa),      bg_, P['bl'], sz=9, h='left')
        wr(ws5, r, 4, f'${capital:,.0f}',  bg_, P['bl'], bold=True, sz=9)
        wr(ws5, r, 5, f'{tasa}%',          bg_, P['am'], sz=9)
        wr(ws5, r, 6, f'{plazo} meses',    bg_, P['gr'], sz=9)
        wr(ws5, r, 7, f'${resultado:,.0f}',bg_, P['cy'], bold=True, sz=10)
        wr(ws5, r, 8, f'+${ganancia:,.0f}',bg_, P['ve'], bold=True, sz=10)
        wr(ws5, r, 9, f'${neto:,.0f}',    bg_, P['cy'], sz=9)
 
    ws5.auto_filter.ref = 'A4:I4'
 
    if sims:
        ws5['K1'] = '#'; ws5['L1'] = 'Capital'; ws5['M1'] = 'Resultado'
        for c in ['K1','L1','M1']:
            ws5[c].fill = fl(P['li']); ws5[c].font = fn(P['bl'], bold=True, sz=9)
        cw(ws5,11,8); cw(ws5,12,16); cw(ws5,13,16)
        rs = 2
        for i3, s in enumerate(sims[:12], 1):
            ws5[f'K{rs}'] = i3
            ws5[f'L{rs}'] = round(float(s.capital_inicial), 2)
            ws5[f'M{rs}'] = round(float(s.resultado_final), 2)
            for c3 in [f'K{rs}',f'L{rs}',f'M{rs}']:
                ws5[c3].fill = fl(P['ca']); ws5[c3].font = fn(P['bl'], sz=9)
            rs += 1
        lc = LineChart(); lc.title = 'Capital vs Resultado'
        lc.style = 10; lc.width = 22; lc.height = 14
        dl = Reference(ws5, min_col=12, max_col=13, min_row=1, max_row=rs-1)
        lc.add_data(dl, titles_from_data=True)
        lc.series[0].graphicalProperties.line.solidFill = P['mo']
        lc.series[1].graphicalProperties.line.solidFill = P['cy']
        lc.series[0].graphicalProperties.line.width = 25000
        lc.series[1].graphicalProperties.line.width = 25000
        ws5.add_chart(lc, 'K4')
 
    # ════════════════════════════════════════════════════
    #  HOJA 6 — PLAN FINANCIERO 50/30/20 + PROYECCIONES
    # ════════════════════════════════════════════════════
    ws6 = wb.create_sheet('Plan Financiero')
    ws6.sheet_properties.tabColor = P['ve']
    cabecera(ws6, 'Plan Financiero Personal',
        f'Regla 50/30/20  ·  Proyecciones de ahorro  ·  {user.nombre}', 6)
 
    ing_base = ing_mens if ing_mens > 0 else (
        ing_t / max(1, len(set(t.fecha.month for t in trans if t.fecha)))
    )
 
    if ing_base > 0:
        n50 = ing_base * 0.5
        n30 = ing_base * 0.3
        n20 = ing_base * 0.2
 
        wr(ws6, 4, 1, '📋  DISTRIBUCIÓN RECOMENDADA (REGLA 50/30/20)',
           P['li'], P['am'], bold=True, sz=11, h='left')
        enc_fila(ws6, 5, [
            ('Categoría',22),('% Asignado',14),('Monto mensual',16),
            ('Monto anual',16),('Descripción',30),('Real actual',14),
        ])
 
        # Agrupa gasto real por bloque 50/30/20
        nec_cats = {'Alimentación','Arriendo','Salud','Transporte','Servicios','Médico'}
        des_cats  = {'Entretenimiento','Ropa','Restaurantes','Tecnología','Viajes','Mascotas','Regalos'}
        real_nec  = sum(gastos_cat.get(c,0) for c in nec_cats)
        real_des  = sum(gastos_cat.get(c,0) for c in des_cats)
 
        plan = [
            ('🟢 Necesidades', '50%', n50, n50*12,
             'Arriendo, comida, transporte, servicios, salud', f'${real_nec:,.0f}', P['ve']),
            ('🟡 Deseos',      '30%', n30, n30*12,
             'Entretenimiento, ropa, salidas, hobbies',        f'${real_des:,.0f}', P['am']),
            ('🔵 Ahorro/Inv.', '20%', n20, n20*12,
             'CDT, metas, pensión voluntaria',                  f'${max(0,bal_t):,.0f}', P['cy']),
        ]
        for i, (cat,pct_,mn_,an_,desc,real,col_) in enumerate(plan, 6):
            bg_ = P['ca'] if i%2==0 else P['bg']; rh(ws6, i, 24)
            wr(ws6,i,1,cat,           bg_,col_,bold=True,sz=10,h='left')
            wr(ws6,i,2,pct_,          bg_,col_,sz=9)
            wr(ws6,i,3,f'${mn_:,.0f}',bg_,P['bl'],bold=True,sz=10)
            wr(ws6,i,4,f'${an_:,.0f}',bg_,P['gr'],sz=9)
            wr(ws6,i,5,desc,          bg_,P['gr'],sz=9,h='left')
            wr(ws6,i,6,real,          bg_,P['cy'],sz=9)
        cw(ws6,1,24); cw(ws6,2,14); cw(ws6,3,16); cw(ws6,4,16); cw(ws6,5,30); cw(ws6,6,14)
 
        # Proyecciones
        wr(ws6, 11, 1, '📈  PROYECCIONES (ahorrando el 20% del ingreso mensual)',
           P['li'], P['cy'], bold=True, sz=11, h='left')
        enc_fila(ws6, 12, [
            ('Horizonte',16),('Ahorro simple',16),('Con 8% EA',16),
            ('Con 10% EA',16),('Con 12% EA',16),('Ventaja c/interés',18),
        ])
 
        def ci(tasa, meses):
            tm = tasa/12; b = 0
            for _ in range(meses): b = b*(1+tm) + n20
            return b
 
        horizontes = [(6,'6 meses'),(12,'1 año'),(24,'2 años'),(36,'3 años'),(60,'5 años'),(120,'10 años')]
        for i,(m,lbl) in enumerate(horizontes, 13):
            simple=n20*m; c8=ci(.08,m); c10=ci(.10,m); c12=ci(.12,m)
            bg_=P['ca'] if i%2==0 else P['bg']; rh(ws6,i,22)
            wr(ws6,i,1,lbl,                bg_,P['bl'],bold=True,sz=9)
            wr(ws6,i,2,f'${simple:,.0f}', bg_,P['gr'],sz=9)
            wr(ws6,i,3,f'${c8:,.0f}',    bg_,P['cy'],sz=9)
            wr(ws6,i,4,f'${c10:,.0f}',   bg_,P['ve'],bold=True,sz=10)
            wr(ws6,i,5,f'${c12:,.0f}',   bg_,P['am'],sz=9)
            wr(ws6,i,6,f'+${c10-simple:,.0f}',bg_,P['ve'],sz=9)
 
        # Gráfico proyecciones
        ws6['H12']='Plazo'; ws6['I12']='Simple'; ws6['J12']='10% EA'
        for c in['H12','I12','J12']:
            ws6[c].fill=fl(P['li']); ws6[c].font=fn(P['bl'],bold=True,sz=9)
        cw(ws6,8,14); cw(ws6,9,14); cw(ws6,10,16)
        for i3,(m,lbl) in enumerate(horizontes,13):
            ws6[f'H{i3}']=lbl
            ws6[f'I{i3}']=round(n20*m,2)
            ws6[f'J{i3}']=round(ci(.10,m),2)
            for c3 in[f'H{i3}',f'I{i3}',f'J{i3}']:
                ws6[c3].fill=fl(P['ca']); ws6[c3].font=fn(P['bl'],sz=9)
        lp=LineChart(); lp.title='Ahorro simple vs interés compuesto'
        lp.style=10; lp.width=22; lp.height=14
        dp=Reference(ws6,min_col=9,max_col=10,min_row=12,max_row=18)
        lp.add_data(dp,titles_from_data=True)
        lp.series[0].graphicalProperties.line.solidFill=P['gr']
        lp.series[1].graphicalProperties.line.solidFill=P['ve']
        lp.series[0].graphicalProperties.line.width=20000
        lp.series[1].graphicalProperties.line.width=28000
        ws6.add_chart(lp,'H16')
    else:
        wr(ws6,5,1,
           'Registra tu ingreso mensual en Mi Perfil para activar el plan personalizado.',
           P['bg'],P['am'],sz=10,h='left')
        cw(ws6,1,60)
 
    # ── Guardar ───────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nombre = f'FinanBot_{user.nombre.replace(" ","_")}_{datetime.now().strftime("%d-%m-%Y")}.xlsx'
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=nombre)
 