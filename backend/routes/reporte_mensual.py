# reporte_mensual.py

import calendar as calendar_lib
import unicodedata
from collections import defaultdict
from datetime import datetime, date
from io import BytesIO

MESES_ES = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
]
DIAS_ES = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']

ICONOS = {
    'Alimentación': '🍔', 'Transporte': '🚌', 'Arriendo': '🏠', 'Salud': '💊',
    'Entretenimiento': '🎬', 'Educación': '📚', 'Ropa': '👗', 'Servicios': '⚡',
    'Mascotas': '🐾', 'Regalos': '🎁', 'Restaurantes': '🍕', 'Viajes': '✈️',
    'Otros gastos': '📦', 'Salario': '💼', 'Freelance': '🧑\u200d💻',
    'Inversión': '📈', 'Negocio': '🏪', 'Regalo': '🎁', 'Otros ingresos': '💵',
}

CATEGORIAS_GASTO_ORDEN = [
    'Alimentación', 'Transporte', 'Arriendo', 'Salud', 'Entretenimiento',
    'Educación', 'Ropa', 'Servicios', 'Mascotas', 'Regalos', 'Viajes', 'Otros gastos'
]
CATEGORIAS_INGRESO_ORDEN = [
    'Salario', 'Freelance', 'Inversión', 'Negocio', 'Regalo', 'Otros ingresos'
]


# ══════════════════════════════════════════════════════════════════
#  HELPERS COMPARTIDOS
# ══════════════════════════════════════════════════════════════════
def _limpiar(texto):
    """Quita tildes/emojis — necesario para el PDF (Helvetica base14 no
    renderiza unicode extendido). Excel sí soporta unicode completo,
    así que ahí se usa el texto tal cual."""
    if not texto:
        return 'Sin categoria'
    nfkd = unicodedata.normalize('NFKD', str(texto))
    ascii_str = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return ''.join(c for c in ascii_str if 32 <= ord(c) <= 126) or 'Sin categoria'


def _get_cat(t):
    """La categoría puede venir como string o como objeto relacionado
    (con .nombre) según el modelo — se soporta ambos casos."""
    cat = getattr(t, 'categoria', None)
    if cat is None:
        return 'Sin categoría'
    if isinstance(cat, str):
        return cat
    nombre = getattr(cat, 'nombre', None)
    return nombre if nombre else str(cat)


def _icono(cat_str):
    return ICONOS.get(cat_str, '💸')


def _fecha_str(fecha):
    return fecha.strftime('%Y-%m-%d') if hasattr(fecha, 'strftime') else str(fecha)


def _semanas_del_mes(anio, mes):
    """Divide el mes en semanas de calendario (lunes a domingo), cada
    una recortada a los límites del mes. Devuelve una lista de tuplas
    (numero_semana, fecha_inicio, fecha_fin)."""
    cal = calendar_lib.Calendar(firstweekday=0)  # lunes = 0
    semanas = []
    num = 1
    for semana in cal.monthdatescalendar(anio, mes):
        dias_del_mes = [d for d in semana if d.month == mes]
        if not dias_del_mes:
            continue
        semanas.append((num, dias_del_mes[0], dias_del_mes[-1]))
        num += 1
    return semanas


def _agrupar_mes(transacciones, anio, mes):
    """Arma todas las agregaciones que necesitan tanto el PDF como el
    Excel, para no repetir la misma lógica dos veces."""
    ingresos = [t for t in transacciones if t.tipo == 'ingreso']
    gastos   = [t for t in transacciones if t.tipo == 'gasto']
    tot_ing  = sum(float(t.monto) for t in ingresos)
    tot_gas  = sum(float(t.monto) for t in gastos)
    balance  = tot_ing - tot_gas

    cat_gastos, cat_ingresos = defaultdict(float), defaultdict(float)
    for t in gastos:
        cat_gastos[_get_cat(t)] += float(t.monto)
    for t in ingresos:
        cat_ingresos[_get_cat(t)] += float(t.monto)

    # Por día (solo días con movimientos)
    por_dia = defaultdict(lambda: {'ingreso': 0.0, 'gasto': 0.0, 'items': []})
    for t in transacciones:
        f = t.fecha if isinstance(t.fecha, date) else datetime.strptime(_fecha_str(t.fecha), '%Y-%m-%d').date()
        por_dia[f][t.tipo] += float(t.monto)
        por_dia[f]['items'].append(t)

    # Por semana (usa los rangos reales de calendario del mes)
    semanas_def = _semanas_del_mes(anio, mes)
    por_semana = []
    for num, inicio, fin in semanas_def:
        ing_sem = gas_sem = 0.0
        for f, datos in por_dia.items():
            if inicio <= f <= fin:
                ing_sem += datos['ingreso']
                gas_sem += datos['gasto']
        por_semana.append({
            'numero': num, 'inicio': inicio, 'fin': fin,
            'ingresos': ing_sem, 'gastos': gas_sem, 'balance': ing_sem - gas_sem,
        })

    ultimo_dia_mes = calendar_lib.monthrange(anio, mes)[1]
    dias_con_datos = len(por_dia)
    promedio_gasto_diario = round(tot_gas / dias_con_datos, 0) if dias_con_datos else 0
    cat_mayor = max(cat_gastos, key=cat_gastos.get) if cat_gastos else None

    return {
        'ingresos': ingresos, 'gastos': gastos,
        'total_ingresos': tot_ing, 'total_gastos': tot_gas, 'balance': balance,
        'cat_gastos': cat_gastos, 'cat_ingresos': cat_ingresos,
        'por_dia': por_dia, 'por_semana': por_semana,
        'ultimo_dia_mes': ultimo_dia_mes, 'dias_con_datos': dias_con_datos,
        'promedio_gasto_diario': promedio_gasto_diario, 'cat_mayor': cat_mayor,
    }


def _barra(pct, largo=20):
    pct = max(0, min(100, pct))
    llenas = int(pct / 100 * largo)
    return '|' * llenas + '.' * (largo - llenas) + f'  {round(pct)}%'


# ══════════════════════════════════════════════════════════════════
#  PDF  ─  mismo tema oscuro que exportar.py, organizado por mes
# ══════════════════════════════════════════════════════════════════
def generar_reporte_pdf_mensual(usuario, transacciones, anio: int, mes: int) -> BytesIO:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, HRFlowable)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    def rgb(r, g, b): return colors.Color(r / 255, g / 255, b / 255)

    BG, CARD, CARD2, LINE, HEADER = rgb(15, 15, 45), rgb(26, 26, 78), rgb(20, 20, 60), rgb(45, 27, 105), rgb(35, 20, 80)
    MORADO, CYAN, VERDE, ROSA, AMAR = rgb(168, 85, 247), rgb(34, 211, 238), rgb(74, 222, 128), rgb(244, 114, 182), rgb(251, 191, 36)
    BLANC, GRIS, MUTED = rgb(255, 255, 255), rgb(196, 181, 253), rgb(107, 114, 128)

    def st(nm, **kw):
        base = {'fontName': 'Helvetica', 'textColor': BLANC, 'leading': 14}
        base.update(kw)
        return ParagraphStyle(nm, **base)

    S_TITULO = st('titulo', fontSize=22, fontName='Helvetica-Bold', textColor=MORADO, leading=26)
    S_SEC    = st('sec', fontSize=12, fontName='Helvetica-Bold', textColor=CYAN, spaceAfter=4, leading=16)
    S_TH     = st('th', fontSize=8, fontName='Helvetica-Bold', textColor=BLANC, alignment=TA_CENTER, leading=11)
    S_TC     = st('tc', fontSize=8, textColor=BLANC, alignment=TA_LEFT, leading=11)
    S_LABEL  = st('lbl', fontSize=8, textColor=GRIS, leading=11)
    S_FOOT   = st('ft', fontSize=7, textColor=MUTED, alignment=TA_CENTER)

    nombre_mes = MESES_ES[mes]
    r = _agrupar_mes(transacciones, anio, mes)
    nombre_pdf = _limpiar(usuario.nombre)

    buf = BytesIO()
    W = A4[0] - 40 * mm
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20 * mm, rightMargin=20 * mm,
                             topMargin=20 * mm, bottomMargin=20 * mm)

    def mk_tabla(hdrs, filas, cws):
        data = [[Paragraph(h, S_TH) for h in hdrs]] + filas
        t = Table(data, colWidths=cws, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HEADER),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [CARD, CARD2]),
            ('GRID', (0, 0), (-1, -1), 0.4, LINE),
            ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, MORADO),
        ]))
        return t

    story = []

    # ── ENCABEZADO ─────────────────────────────────────────
    fecha_str = datetime.now().strftime('%d/%m/%Y  %H:%M')
    detalles = Table([
        [Paragraph(f'<b>{nombre_pdf}</b>', st('un', fontSize=11, fontName='Helvetica-Bold'))],
        [Paragraph(_limpiar(usuario.correo), st('uc', fontSize=9, textColor=GRIS))],
        [Paragraph(fecha_str, st('ud', fontSize=9, textColor=MUTED))],
        [Paragraph(f'Reporte mensual — {nombre_mes} {anio}', st('ur', fontSize=8, textColor=MORADO))],
    ], colWidths=[W * 0.6])
    detalles.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 1), ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]))
    hd = Table([[Paragraph(f'FinanBot — {nombre_mes} {anio}', S_TITULO), detalles]], colWidths=[W * 0.55, W * 0.45])
    hd.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), CARD),
        ('TOPPADDING', (0, 0), (-1, -1), 14), ('BOTTOMPADDING', (0, 0), (-1, -1), 14),
        ('LEFTPADDING', (0, 0), (-1, -1), 18), ('RIGHTPADDING', (0, 0), (-1, -1), 18),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, -1), 2, MORADO), ('LINEABOVE', (0, 0), (-1, 0), 3, CYAN),
    ]))
    story += [hd, Spacer(1, 16)]

    # ── KPIs DEL MES ────────────────────────────────────────
    story += [Paragraph(f'Resumen de {nombre_mes}', S_SEC), Spacer(1, 6)]

    def tarjeta(label, valor, color, sub=''):
        content = [[Paragraph(label, S_LABEL)], [Paragraph(valor, st('tv', fontSize=15, fontName='Helvetica-Bold', textColor=color, leading=18))]]
        if sub:
            content.append([Paragraph(sub, st('ts', fontSize=7, textColor=MUTED, leading=10))])
        inn = Table(content, colWidths=[W / 5 - 8])
        inn.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), CARD),
            ('TOPPADDING', (0, 0), (-1, -1), 9), ('BOTTOMPADDING', (0, 0), (-1, -1), 9),
            ('LEFTPADDING', (0, 0), (-1, -1), 10), ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('LINEABOVE', (0, 0), (-1, 0), 3, color),
        ]))
        return inn

    bal_col = VERDE if r['balance'] >= 0 else ROSA
    tarjetas = Table([[
        tarjeta('INGRESOS DEL MES', f"${r['total_ingresos']:,.0f}", CYAN, f"{len(r['ingresos'])} registros"),
        tarjeta('GASTOS DEL MES', f"${r['total_gastos']:,.0f}", MORADO, f"{len(r['gastos'])} registros"),
        tarjeta('BALANCE DEL MES', f"${r['balance']:,.0f}", bal_col, 'Positivo' if r['balance'] >= 0 else 'Negativo'),
        tarjeta('PROMEDIO GASTO/DÍA', f"${r['promedio_gasto_diario']:,.0f}", AMAR, f"{r['dias_con_datos']} días con datos"),
        tarjeta('DÍAS DEL MES', f"{r['dias_con_datos']}/{r['ultimo_dia_mes']}", GRIS, 'con movimientos'),
    ]], colWidths=[W / 5 - 4] * 5, hAlign='LEFT')
    tarjetas.setStyle(TableStyle([('LEFTPADDING', (0, 0), (-1, -1), 4), ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                                   ('TOPPADDING', (0, 0), (-1, -1), 0), ('BOTTOMPADDING', (0, 0), (-1, -1), 0)]))
    story += [tarjetas, Spacer(1, 18)]

    # ── RESUMEN SEMANA A SEMANA ────────────────────────────
    if r['por_semana']:
        story += [Paragraph('Resumen semana a semana', S_SEC), Spacer(1, 6)]
        filas_sem = []
        for s in r['por_semana']:
            rango = f"{s['inicio'].strftime('%d/%m')} – {s['fin'].strftime('%d/%m')}"
            col_bal = VERDE if s['balance'] >= 0 else ROSA
            filas_sem.append([
                Paragraph(f"Semana {s['numero']}", S_TC),
                Paragraph(rango, st('rg', fontSize=8, textColor=GRIS, alignment=TA_CENTER)),
                Paragraph(f"${s['ingresos']:,.0f}", st('si', fontSize=8, textColor=CYAN, alignment=TA_CENTER)),
                Paragraph(f"${s['gastos']:,.0f}", st('sg', fontSize=8, textColor=MORADO, alignment=TA_CENTER)),
                Paragraph(f"${s['balance']:,.0f}", st('sb', fontSize=8, fontName='Helvetica-Bold', textColor=col_bal, alignment=TA_CENTER)),
            ])
        story += [mk_tabla(['Semana', 'Rango', 'Ingresos', 'Gastos', 'Balance'], filas_sem,
                            [W * 0.16, W * 0.26, W * 0.19, W * 0.19, W * 0.20]), Spacer(1, 18)]

    # ── GASTOS POR CATEGORÍA ────────────────────────────────
    if r['cat_gastos']:
        story += [Paragraph('Gastos por categoría', S_SEC), Spacer(1, 6)]
        filas_cat = []
        for cat in sorted(r['cat_gastos'], key=r['cat_gastos'].get, reverse=True):
            monto = r['cat_gastos'][cat]
            pct = round(monto / r['total_gastos'] * 100) if r['total_gastos'] else 0
            col = ROSA if pct > 40 else AMAR if pct > 25 else VERDE
            filas_cat.append([
                Paragraph(_limpiar(cat), S_TC),
                Paragraph(f'${monto:,.0f}', st('mc', fontSize=8, alignment=TA_CENTER)),
                Paragraph(f'{pct}%', st('pc', fontSize=8, fontName='Helvetica-Bold', textColor=col, alignment=TA_CENTER)),
                Paragraph(_barra(pct), st('bc', fontSize=7, textColor=col, fontName='Courier')),
            ])
        story += [mk_tabla(['Categoría', 'Monto', '%', 'Distribución'], filas_cat,
                            [W * 0.26, W * 0.17, W * 0.10, W * 0.47]), Spacer(1, 18)]

    # ── INGRESOS POR CATEGORÍA ──────────────────────────────
    if r['cat_ingresos']:
        story += [Paragraph('Ingresos por categoría', S_SEC), Spacer(1, 6)]
        filas_ing = []
        for cat in sorted(r['cat_ingresos'], key=r['cat_ingresos'].get, reverse=True):
            monto = r['cat_ingresos'][cat]
            pct = round(monto / r['total_ingresos'] * 100) if r['total_ingresos'] else 0
            filas_ing.append([
                Paragraph(_limpiar(cat), S_TC),
                Paragraph(f'${monto:,.0f}', st('mci', fontSize=8, textColor=VERDE, alignment=TA_CENTER)),
                Paragraph(f'{pct}%', st('pci', fontSize=8, textColor=CYAN, alignment=TA_CENTER)),
            ])
        story += [mk_tabla(['Fuente', 'Monto', '%'], filas_ing, [W * 0.5, W * 0.3, W * 0.2]), Spacer(1, 18)]

    # ── DETALLE DÍA A DÍA ────────────────────────────────────
    if r['por_dia']:
        story += [Paragraph('Detalle día a día', S_SEC), Spacer(1, 6)]
        filas_dia = []
        saldo_acum = 0.0
        for f in sorted(r['por_dia']):
            datos = r['por_dia'][f]
            saldo_acum += datos['ingreso'] - datos['gasto']
            nombre_dia = DIAS_ES[f.weekday()]
            filas_dia.append([
                Paragraph(f.strftime('%d/%m'), st('fd', fontSize=8, alignment=TA_CENTER)),
                Paragraph(nombre_dia, st('nd', fontSize=8, textColor=GRIS, alignment=TA_CENTER)),
                Paragraph(f"${datos['ingreso']:,.0f}" if datos['ingreso'] else '—', st('id2', fontSize=8, textColor=VERDE, alignment=TA_CENTER)),
                Paragraph(f"${datos['gasto']:,.0f}" if datos['gasto'] else '—', st('gd', fontSize=8, textColor=ROSA, alignment=TA_CENTER)),
                Paragraph(f'${saldo_acum:,.0f}', st('sa', fontSize=8, fontName='Helvetica-Bold',
                          textColor=CYAN if saldo_acum >= 0 else ROSA, alignment=TA_CENTER)),
            ])
        story += [mk_tabla(['Fecha', 'Día', 'Ingresos', 'Gastos', 'Saldo acum.'], filas_dia,
                            [W * 0.16, W * 0.24, W * 0.20, W * 0.20, W * 0.20]), Spacer(1, 18)]

    # ── TODAS LAS TRANSACCIONES DEL MES ─────────────────────
    if transacciones:
        story += [Paragraph('Transacciones del mes', S_SEC), Spacer(1, 6)]
        filas_tr = []
        for i, t in enumerate(sorted(transacciones, key=lambda x: x.fecha), 1):
            col_tipo = VERDE if t.tipo == 'ingreso' else ROSA
            sgn = '+' if t.tipo == 'ingreso' else '-'
            fd = t.fecha.strftime('%d/%m') if hasattr(t.fecha, 'strftime') else str(t.fecha)
            filas_tr.append([
                Paragraph(str(i), st('fn', fontSize=8, textColor=MUTED, alignment=TA_CENTER)),
                Paragraph(fd, st('fdt', fontSize=8, alignment=TA_CENTER)),
                Paragraph(t.tipo.capitalize(), st('ft2', fontSize=8, textColor=col_tipo, fontName='Helvetica-Bold', alignment=TA_CENTER)),
                Paragraph(_limpiar(_get_cat(t)), S_TC),
                Paragraph(_limpiar(t.descripcion) if t.descripcion else '—', st('fd2', fontSize=8, textColor=GRIS)),
                Paragraph(f'{sgn}${float(t.monto):,.0f}', st('fm', fontSize=8, fontName='Helvetica-Bold', textColor=col_tipo, alignment=TA_RIGHT)),
            ])
        story += [mk_tabla(['#', 'Fecha', 'Tipo', 'Categoría', 'Descripción', 'Monto'], filas_tr,
                            [W * 0.05, W * 0.11, W * 0.11, W * 0.19, W * 0.34, W * 0.20]), Spacer(1, 18)]

    # ── PIE DE PÁGINA ────────────────────────────────────────
    story += [HRFlowable(width=W, thickness=0.5, color=LINE), Spacer(1, 6),
              Paragraph(f'FinanBot  |  {nombre_pdf}  |  {nombre_mes} {anio}  |  Documento confidencial', S_FOOT)]

    def fondo_pagina(canvas, doc_):
        canvas.saveState()
        canvas.setFillColor(BG); canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
        canvas.setFillColor(MORADO); canvas.rect(0, A4[1] - 5, A4[0] * 0.5, 5, fill=1, stroke=0)
        canvas.setFillColor(CYAN); canvas.rect(A4[0] * 0.5, A4[1] - 5, A4[0] * 0.5, 5, fill=1, stroke=0)
        canvas.setFillColor(MUTED); canvas.setFont('Helvetica', 7)
        canvas.drawRightString(A4[0] - 20 * mm, 12 * mm, f'Pág. {doc_.page}')
        canvas.restoreState()

    doc.build(story, onFirstPage=fondo_pagina, onLaterPages=fondo_pagina)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════
#  EXCEL  ─  mismo tema oscuro que excel.py, con hojas propias del mes
# ══════════════════════════════════════════════════════════════════
def generar_reporte_excel_mensual(usuario, transacciones, anio: int, mes: int) -> BytesIO:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter

    nombre_mes = MESES_ES[mes]
    r = _agrupar_mes(transacciones, anio, mes)

    P = {
        'mo': 'C026D3', 'md': '7E22CE', 'cy': '06B6D4', 've': '22C55E',
        'ro': 'F472B6', 'am': 'F59E0B', 're': 'EF4444', 'bg': '0F0F2D',
        'ca': '1A1A4E', 'li': '2D1B69', 'bl': 'FFFFFF', 'gr': '9CA3AF', 'mu': '6B7280',
    }

    wb = openpyxl.Workbook()

    def fl(h): return PatternFill('solid', fgColor=h)
    def fn(h=None, bold=False, sz=10, italic=False):
        return Font(color=h or P['bl'], bold=bold, size=sz, italic=italic, name='Segoe UI')
    def al(h='center', v='center', wrap=False): return Alignment(horizontal=h, vertical=v, wrapText=wrap)
    def bd(col=None):
        s = Side(style='thin', color=col or P['li'])
        return Border(left=s, right=s, top=s, bottom=s)
    def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
    def rh(ws, row, h): ws.row_dimensions[row].height = h

    def wr(ws, row, col, val, bg=None, fg=None, bold=False, sz=10, h='center', brd=True, italic=False):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fl(bg or P['ca']); c.font = fn(fg or P['bl'], bold, sz, italic)
        c.alignment = al(h)
        if brd: c.border = bd()
        return c

    def fondo_hoja(ws, filas=200, cols=16):
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows(min_row=1, max_row=filas, min_col=1, max_col=cols):
            for c in row: c.fill = fl(P['bg'])

    def cabecera(ws, titulo, sub, cols=8):
        fondo_hoja(ws)
        rh(ws, 1, 46); rh(ws, 2, 20)
        ws.merge_cells(f'A1:{get_column_letter(cols)}1')
        c1 = ws['A1']; c1.value = titulo; c1.fill = fl(P['mo']); c1.font = fn(P['bl'], bold=True, sz=18); c1.alignment = al()
        ws.merge_cells(f'A2:{get_column_letter(cols)}2')
        c2 = ws['A2']; c2.value = sub; c2.fill = fl(P['md']); c2.font = fn(P['gr'], sz=9, italic=True); c2.alignment = al()

    def enc_fila(ws, row, defs, bg=None):
        rh(ws, row, 26)
        for i, (txt, ancho) in enumerate(defs, 1):
            c = ws.cell(row=row, column=i, value=txt)
            c.fill = fl(bg or P['li']); c.font = fn(P['bl'], bold=True, sz=9)
            c.alignment = al(); c.border = bd(); cw(ws, i, ancho)

    def kpi_block(ws, row_start, items):
        rh(ws, row_start, 18); rh(ws, row_start + 1, 36); rh(ws, row_start + 2, 12)
        for i, (lbl, val, col) in enumerate(items, 1):
            for r_off, tx, sz_, bld in [(0, lbl, 8, False), (1, val, 14, True), (2, '', 7, False)]:
                c = ws.cell(row=row_start + r_off, column=i, value=tx)
                c.fill = fl(P['bg'] if r_off == 0 else P['ca'])
                c.font = fn(P['mu'] if r_off == 0 else col, bld, sz_)
                c.alignment = al(); c.border = bd(col)
            cw(ws, i, 18)

    # ════════════════════════════════════════════════════
    #  HOJA 1 — RESUMEN DEL MES
    # ════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = 'Resumen del mes'
    ws1.sheet_properties.tabColor = P['mo']
    cabecera(ws1, f'FinanBot — {nombre_mes} {anio}',
             f'Usuario: {usuario.nombre}  ·  {usuario.correo}  ·  Generado {datetime.now().strftime("%d/%m/%Y %H:%M")}', 7)

    kpi_block(ws1, 4, [
        ('INGRESOS DEL MES', f"${r['total_ingresos']:,.0f}", P['cy']),
        ('GASTOS DEL MES',   f"${r['total_gastos']:,.0f}",   P['mo']),
        ('BALANCE',          f"${r['balance']:,.0f}",        P['ve'] if r['balance'] >= 0 else P['re']),
        ('PROM. GASTO/DÍA',  f"${r['promedio_gasto_diario']:,.0f}", P['am']),
        ('DÍAS CON DATOS',   f"{r['dias_con_datos']}/{r['ultimo_dia_mes']}", P['gr']),
    ])

    rh(ws1, 7, 5)
    for i in range(1, 8):
        ws1.cell(row=7, column=i).fill = fl(P['li'])

    # Gráfico: ingresos vs gastos por semana
    enc_fila(ws1, 8, [('Semana', 14), ('Rango', 18), ('Ingresos', 14), ('Gastos', 14), ('Balance', 14)])
    for i, s in enumerate(r['por_semana'], 9):
        bg_ = P['ca'] if i % 2 == 0 else P['bg']
        col_bal = P['ve'] if s['balance'] >= 0 else P['re']
        rh(ws1, i, 20)
        wr(ws1, i, 1, f"Semana {s['numero']}", bg_, P['bl'], sz=9)
        wr(ws1, i, 2, f"{s['inicio'].strftime('%d/%m')} - {s['fin'].strftime('%d/%m')}", bg_, P['gr'], sz=9)
        wr(ws1, i, 3, round(s['ingresos'], 2), bg_, P['cy'], sz=9)
        wr(ws1, i, 4, round(s['gastos'], 2), bg_, P['mo'], sz=9)
        wr(ws1, i, 5, round(s['balance'], 2), bg_, col_bal, bold=True, sz=9)

    fila_grafico = 9 + len(r['por_semana']) + 1
    if r['por_semana']:
        bar = BarChart(); bar.type = 'col'; bar.title = 'Ingresos vs Gastos por semana'
        bar.style = 10; bar.width = 22; bar.height = 13
        data_ref = Reference(ws1, min_col=3, max_col=4, min_row=8, max_row=8 + len(r['por_semana']))
        cat_ref = Reference(ws1, min_col=1, min_row=9, max_row=8 + len(r['por_semana']))
        bar.add_data(data_ref, titles_from_data=True); bar.set_categories(cat_ref)
        bar.series[0].graphicalProperties.solidFill = P['cy']
        bar.series[1].graphicalProperties.solidFill = P['mo']
        ws1.add_chart(bar, f'A{fila_grafico}')

    # ════════════════════════════════════════════════════
    #  HOJA 2 — DÍA A DÍA
    # ════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Día a día')
    ws2.sheet_properties.tabColor = P['cy']
    cabecera(ws2, f'Detalle día a día — {nombre_mes} {anio}',
             f"{r['dias_con_datos']} días con movimientos de {r['ultimo_dia_mes']} en el mes", 6)
    enc_fila(ws2, 4, [('Fecha', 13), ('Día', 14), ('Ingresos', 14), ('Gastos', 14), ('Balance del día', 15), ('Saldo acumulado', 16)])

    saldo_acum = 0.0
    for i, f in enumerate(sorted(r['por_dia']), 5):
        datos = r['por_dia'][f]
        bal_dia = datos['ingreso'] - datos['gasto']
        saldo_acum += bal_dia
        bg_ = P['ca'] if i % 2 == 0 else P['bg']
        rh(ws2, i, 19)
        wr(ws2, i, 1, f.strftime('%d/%m/%Y'), bg_, P['gr'], sz=9)
        wr(ws2, i, 2, DIAS_ES[f.weekday()], bg_, P['gr'], sz=9)
        wr(ws2, i, 3, round(datos['ingreso'], 2), bg_, P['ve'], sz=9)
        wr(ws2, i, 4, round(datos['gasto'], 2), bg_, P['ro'], sz=9)
        wr(ws2, i, 5, round(bal_dia, 2), bg_, P['ve'] if bal_dia >= 0 else P['re'], bold=True, sz=9)
        wr(ws2, i, 6, round(saldo_acum, 2), bg_, P['cy'] if saldo_acum >= 0 else P['re'], sz=9)

    if r['por_dia']:
        fila_fin = 5 + len(r['por_dia'])
        line_ref_row = fila_fin
        bar2 = BarChart(); bar2.type = 'col'; bar2.title = 'Balance diario'
        bar2.style = 10; bar2.width = 24; bar2.height = 13
        d2 = Reference(ws2, min_col=5, min_row=4, max_row=line_ref_row - 1)
        c2 = Reference(ws2, min_col=1, min_row=5, max_row=line_ref_row - 1)
        bar2.add_data(d2, titles_from_data=True); bar2.set_categories(c2)
        bar2.series[0].graphicalProperties.solidFill = P['ve']
        ws2.add_chart(bar2, f'H4')

    # ════════════════════════════════════════════════════
    #  HOJA 3 — CATEGORÍAS DEL MES
    # ════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Categorías del mes')
    ws3.sheet_properties.tabColor = P['am']
    cabecera(ws3, f'Categorías — {nombre_mes} {anio}',
             f"{len(r['cat_gastos'])} categorías de gasto  ·  {len(r['cat_ingresos'])} fuentes de ingreso", 6)

    wr(ws3, 4, 1, '💸  GASTOS POR CATEGORÍA', P['li'], P['mo'], bold=True, sz=11, h='left')
    enc_fila(ws3, 5, [('Categoría', 24), ('Monto', 15), ('% del total', 13), ('Distribución', 20)], P['re'])
    fila = 6
    for cat in CATEGORIAS_GASTO_ORDEN:
        if cat not in r['cat_gastos']:
            continue
        monto = r['cat_gastos'][cat]
        pct = round(monto / r['total_gastos'] * 100, 1) if r['total_gastos'] else 0
        bg_ = P['ca'] if fila % 2 == 0 else P['bg']
        col_ = P['re'] if pct > 40 else P['am'] if pct > 25 else P['ve']
        rh(ws3, fila, 19)
        wr(ws3, fila, 1, f'{_icono(cat)} {cat}', bg_, P['bl'], sz=9, h='left')
        wr(ws3, fila, 2, round(monto, 2), bg_, P['ro'], bold=True, sz=9)
        wr(ws3, fila, 3, f'{pct}%', bg_, col_, sz=9)
        wr(ws3, fila, 4, _barra(pct, 15), bg_, col_, sz=8)
        fila += 1

    fila_ing_titulo = fila + 1
    wr(ws3, fila_ing_titulo, 1, '💰  INGRESOS POR FUENTE', P['li'], P['cy'], bold=True, sz=11, h='left')
    enc_fila(ws3, fila_ing_titulo + 1, [('Fuente', 24), ('Monto', 15), ('% del total', 13)], P['ve'])
    fila2 = fila_ing_titulo + 2
    for cat in CATEGORIAS_INGRESO_ORDEN:
        if cat not in r['cat_ingresos']:
            continue
        monto = r['cat_ingresos'][cat]
        pct = round(monto / r['total_ingresos'] * 100, 1) if r['total_ingresos'] else 0
        bg_ = P['ca'] if fila2 % 2 == 0 else P['bg']
        rh(ws3, fila2, 19)
        wr(ws3, fila2, 1, f'{_icono(cat)} {cat}', bg_, P['bl'], sz=9, h='left')
        wr(ws3, fila2, 2, round(monto, 2), bg_, P['ve'], bold=True, sz=9)
        wr(ws3, fila2, 3, f'{pct}%', bg_, P['cy'], sz=9)
        fila2 += 1

    if r['cat_gastos']:
        ws3['H1'] = 'Categoría'; ws3['I1'] = 'Monto'
        ws3['H1'].fill = fl(P['li']); ws3['H1'].font = fn(P['bl'], bold=True, sz=9)
        ws3['I1'].fill = fl(P['li']); ws3['I1'].font = fn(P['bl'], bold=True, sz=9)
        cw(ws3, 8, 24); cw(ws3, 9, 14)
        rp = 2
        for cat, monto in sorted(r['cat_gastos'].items(), key=lambda x: x[1], reverse=True):
            ws3[f'H{rp}'] = f'{_icono(cat)} {cat}'
            ws3[f'I{rp}'] = round(monto, 2)
            ws3[f'H{rp}'].fill = fl(P['ca']); ws3[f'H{rp}'].font = fn(P['bl'], sz=9)
            ws3[f'I{rp}'].fill = fl(P['ca']); ws3[f'I{rp}'].font = fn(P['am'], sz=9)
            rp += 1
        pie = PieChart(); pie.title = f'Distribución de gastos — {nombre_mes}'
        pie.style = 10; pie.width = 18; pie.height = 13
        pd_ = Reference(ws3, min_col=9, min_row=1, max_row=rp - 1)
        pc_ = Reference(ws3, min_col=8, min_row=2, max_row=rp - 1)
        pie.add_data(pd_, titles_from_data=True); pie.set_categories(pc_)
        ws3.add_chart(pie, 'K5')

    # ════════════════════════════════════════════════════
    #  HOJA 4 — TRANSACCIONES DEL MES
    # ════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Transacciones')
    ws4.sheet_properties.tabColor = P['ve']
    cabecera(ws4, f'Transacciones — {nombre_mes} {anio}',
             f"{len(transacciones)} registros  ·  Ingresos: ${r['total_ingresos']:,.0f}  ·  Gastos: ${r['total_gastos']:,.0f}", 8)
    enc_fila(ws4, 4, [('#', 5), ('Fecha', 13), ('Tipo', 11), ('Categoría', 22), ('Descripción', 30), ('Monto', 15), ('Saldo acum.', 16)])

    saldo = 0.0
    for idx, t in enumerate(sorted(transacciones, key=lambda x: x.fecha), 1):
        cat_t = _get_cat(t)
        mn = float(t.monto)
        saldo += mn if t.tipo == 'ingreso' else -mn
        row = idx + 4
        bg_ = P['ca'] if idx % 2 == 0 else P['bg']
        col_ = P['ve'] if t.tipo == 'ingreso' else P['ro']
        sgn = '+' if t.tipo == 'ingreso' else '-'
        fd = t.fecha.strftime('%d/%m/%Y') if hasattr(t.fecha, 'strftime') else str(t.fecha)
        rh(ws4, row, 19)
        wr(ws4, row, 1, idx, bg_, P['mu'], sz=9)
        wr(ws4, row, 2, fd, bg_, P['gr'], sz=9)
        wr(ws4, row, 3, t.tipo.capitalize(), bg_, col_, bold=True, sz=9)
        wr(ws4, row, 4, f'{_icono(cat_t)} {cat_t}', bg_, P['bl'], sz=9, h='left')
        wr(ws4, row, 5, t.descripcion or '—', bg_, P['gr'], sz=9, h='left')
        wr(ws4, row, 6, f'{sgn}${mn:,.0f}', bg_, col_, bold=True, sz=9)
        wr(ws4, row, 7, f'${saldo:,.0f}', bg_, P['cy'] if saldo >= 0 else P['re'], sz=9)

    if transacciones:
        ws4.auto_filter.ref = f'A4:G{4 + len(transacciones)}'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output