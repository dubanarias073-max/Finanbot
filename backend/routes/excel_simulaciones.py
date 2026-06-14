# routes/excel_simulaciones.py
# Exporta una simulación de inversión a Excel con tema oscuro morado/cyan
from flask import Blueprint, request, send_file, jsonify
from datetime import datetime
import io

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

excel_sim_bp = Blueprint('excel_simulaciones', __name__)


# ─────────────────────────────────────────────────────────
#  POST /api/simulaciones/exportar
#  Recibe los datos de la simulación y devuelve un .xlsx
# ─────────────────────────────────────────────────────────
@excel_sim_bp.route('/exportar', methods=['POST'])
def exportar_simulacion_excel():
    data = request.get_json()
    if not data:
        return jsonify({'mensaje': '❌ No se recibieron datos'}), 400

    # ── Validar campos ────────────────────────────────────
    campos = ['capital_inicial', 'tasa_retorno', 'plazo_meses']
    for campo in campos:
        if data.get(campo) is None:
            return jsonify({'mensaje': f'❌ Falta el campo: {campo}'}), 400

    try:
        capital        = float(data['capital_inicial'])
        tasa           = float(data['tasa_retorno'])
        plazo          = int(data['plazo_meses'])
        aporte_mensual = float(data.get('aporte_mensual', 0))
        nombre_tipo    = str(data.get('nombre_tipo', 'Inversión'))
    except (ValueError, TypeError):
        return jsonify({'mensaje': '❌ Los valores deben ser numéricos'}), 400

    # ── Cálculo completo ──────────────────────────────────
    tasa_mensual   = tasa / 100 / 12
    total_invertido = capital + aporte_mensual * plazo

    # Proyección mes a mes
    proyeccion = []
    balance    = capital
    for i in range(plazo + 1):
        cap_acum  = capital + aporte_mensual * i
        int_acum  = max(round(balance - cap_acum, 2), 0)
        proyeccion.append({
            'mes':       i,
            'periodo':   'Inicio' if i == 0 else f'Mes {i}',
            'valor':     round(balance, 2),
            'capital':   round(cap_acum, 2),
            'intereses': int_acum,
        })
        if i < plazo:
            balance = balance * (1 + tasa_mensual) + aporte_mensual

    valor_final  = round(balance, 2)
    ganancia     = round(max(valor_final - total_invertido, 0), 2)
    retencion    = round(ganancia * 0.04, 2)
    saldo_neto   = round(valor_final - retencion, 2)
    rentabilidad = round(((valor_final / capital) - 1) * 100, 2) if capital > 0 else 0
    pct_ganancia = round(((valor_final / total_invertido) - 1) * 100, 2) if total_invertido > 0 else 0

    # Comparador de escenarios
    escenarios = []
    for e_tasa, e_nombre, e_emoji in [
        (5,  'CDT Básico',         '🏦'),
        (8,  'CDT Premium',        '💎'),
        (10, 'Fondo de inversión', '📊'),
        (15, 'Acciones BVC',       '📈'),
        (20, 'Startups',           '🚀'),
    ]:
        tm = e_tasa / 100 / 12
        b  = capital
        for _ in range(plazo):
            b = b * (1 + tm) + aporte_mensual
        b = round(b, 2)
        g = round(max(b - total_invertido, 0), 2)
        r = round(((b / capital) - 1) * 100, 2) if capital > 0 else 0
        escenarios.append({
            'nombre':      f'{e_emoji} {e_nombre}',
            'tasa':        e_tasa,
            'valor_final': b,
            'ganancia':    g,
            'rentabilidad': r,
            'es_actual':   abs(e_tasa - tasa) < 0.01,
        })

    # Texto del plazo
    if plazo < 12:
        plazo_texto = f'{plazo} mes{"es" if plazo > 1 else ""}'
    elif plazo == 12:
        plazo_texto = '1 año'
    elif plazo % 12 == 0:
        a = plazo // 12
        plazo_texto = f'{a} año{"s" if a > 1 else ""}'
    else:
        a  = plazo // 12
        mr = plazo % 12
        plazo_texto = f'{a} año{"s" if a > 1 else ""} y {mr} mes{"es" if mr > 1 else ""}'

    # ── PALETA OSCURA ─────────────────────────────────────
    P = {
        'mo': 'C026D3',  # morado principal
        'md': '7E22CE',  # morado oscuro
        'mp': 'A855F7',  # morado pastel
        'cy': '06B6D4',  # cyan
        'ce': '22D3EE',  # cyan claro
        've': '22C55E',  # verde
        'vl': '4ADE80',  # verde claro
        'ro': 'F472B6',  # rosa
        'am': 'F59E0B',  # amarillo
        'ac': 'FBBF24',  # amarillo claro
        're': 'EF4444',  # rojo
        'bg': '0F0F2D',  # fondo página
        'ca': '1A1A4E',  # fondo card
        'c2': '14143C',  # fondo card alt
        'li': '2D1B69',  # líneas/bordes
        'lc': '3D2B79',  # líneas claras
        'bl': 'FFFFFF',  # blanco
        'gr': '9CA3AF',  # gris secundario
        'gd': 'C4B5FD',  # gris violeta (texto label)
        'mu': '6B7280',  # muted
    }

    wb = openpyxl.Workbook()

    # ── HELPERS ───────────────────────────────────────────
    def fl(h):
        return PatternFill('solid', fgColor=h)

    def fn(h=None, bold=False, sz=10, italic=False, name='Segoe UI'):
        return Font(color=h or P['bl'], bold=bold, size=sz, italic=italic, name=name)

    def al(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrapText=wrap)

    def bd(col=None, style='thin'):
        s = Side(style=style, color=col or P['li'])
        return Border(left=s, right=s, top=s, bottom=s)

    def bd_bottom(col=None, sz='medium'):
        return Border(bottom=Side(style=sz, color=col or P['mo']))

    def cw(ws, col, w):
        ws.column_dimensions[get_column_letter(col)].width = w

    def rh(ws, row, h):
        ws.row_dimensions[row].height = h

    def cel(ws, row, col, val='', bg=None, fg=None, bold=False, sz=10,
            h='center', v='center', wrap=False, brd=True, italic=False):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = fl(bg or P['ca'])
        c.font      = fn(fg or P['bl'], bold, sz, italic)
        c.alignment = al(h, v, wrap)
        if brd:
            c.border = bd()
        return c

    def fondo(ws, filas=200, cols=16):
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows(min_row=1, max_row=filas, min_col=1, max_col=cols):
            for c in row:
                c.fill = fl(P['bg'])

    def merge_cel(ws, r1, c1, r2, c2, val='', bg=None, fg=None,
                  bold=False, sz=11, h='center', wrap=False, italic=False):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        c = ws.cell(row=r1, column=c1, value=val)
        c.fill      = fl(bg or P['ca'])
        c.font      = fn(fg or P['bl'], bold, sz, italic)
        c.alignment = al(h, 'center', wrap)
        return c

    def barra_pct(pct, largo=20):
        """Barra visual ASCII proporcional al porcentaje."""
        lleno = max(0, min(int(pct / 5), largo))
        vacio = largo - lleno
        return '█' * lleno + '░' * vacio + f'  {pct:.1f}%'

    def fila_enc(ws, row, defs, bg=None, fg=None):
        """Dibuja fila de encabezado de tabla."""
        rh(ws, row, 28)
        for i, (txt, ancho) in enumerate(defs, 1):
            c = ws.cell(row=row, column=i, value=txt)
            c.fill      = fl(bg or P['li'])
            c.font      = fn(fg or P['bl'], bold=True, sz=9)
            c.alignment = al()
            c.border    = bd(P['mp'])
            cw(ws, i, ancho)


    # ════════════════════════════════════════════════════════
    #  HOJA 1 — PORTADA / RESUMEN EJECUTIVO
    # ════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = '📊 Resumen'
    ws1.sheet_properties.tabColor = P['mo']
    fondo(ws1, 60, 12)

    # ── Barra de título ──────────────────────────────────
    rh(ws1, 1, 56)
    merge_cel(ws1, 1, 1, 1, 8,
              val='FinanBot — Simulación de Inversión',
              bg=P['mo'], fg=P['bl'], bold=True, sz=20)

    rh(ws1, 2, 22)
    merge_cel(ws1, 2, 1, 2, 8,
              val=f'{nombre_tipo}  ·  {plazo_texto}  ·  Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
              bg=P['md'], fg=P['gd'], sz=9, italic=True)

    # ── Línea separadora ─────────────────────────────────
    rh(ws1, 3, 6)
    for i in range(1, 9):
        c = ws1.cell(row=3, column=i)
        c.fill = fl(P['mp'] if i <= 4 else P['cy'])

    # ── KPIs principales (fila 5–7) ──────────────────────
    rh(ws1, 4, 10)
    rh(ws1, 5, 42)
    rh(ws1, 6, 18)

    kpis = [
        ('CAPITAL INICIAL',   f'${capital:,.0f}',        P['ce'],  1),
        ('VALOR FINAL',       f'${valor_final:,.0f}',    P['ac'],  2),
        ('GANANCIA BRUTA',    f'${ganancia:,.0f}',       P['vl'],  3),
        ('SALDO NETO',        f'${saldo_neto:,.0f}',     P['ce'],  4),
        ('TASA ANUAL',        f'{tasa}%',                 P['mp'],  5),
        ('RENTABILIDAD',      f'{rentabilidad:.2f}%',    P['vl'],  6),
        ('RETENCIÓN (4%)',    f'${retencion:,.0f}',      P['ro'],  7),
        ('PLAZO',             plazo_texto,                P['gr'],  8),
    ]

    for label, val, color, col in kpis:
        # Label
        c_lbl = ws1.cell(row=4, column=col, value=label)
        c_lbl.fill = fl(P['bg']); c_lbl.font = fn(P['mu'], sz=8)
        c_lbl.alignment = al(); c_lbl.border = bd(color)
        # Valor
        c_val = ws1.cell(row=5, column=col, value=val)
        c_val.fill = fl(P['ca']); c_val.font = fn(color, bold=True, sz=15)
        c_val.alignment = al(); c_val.border = bd(color)
        # Fila inferior decorativa
        c_dec = ws1.cell(row=6, column=col)
        c_dec.fill = fl(P['ca']); c_dec.border = bd(color)
        cw(ws1, col, 17)

    # ── Sección parámetros ───────────────────────────────
    rh(ws1, 8, 8)
    merge_cel(ws1, 8, 1, 8, 4,
              val='⚙️  PARÁMETROS DE ENTRADA',
              bg=P['li'], fg=P['gd'], bold=True, sz=11, h='left')

    params = [
        ('Tipo de inversión',     nombre_tipo),
        ('Capital inicial',       f'${capital:,.2f}'),
        ('Aporte mensual',        f'${aporte_mensual:,.2f}'),
        ('Tasa anual efectiva',   f'{tasa}%'),
        ('Tasa mensual efectiva', f'{tasa/12:.4f}%'),
        ('Plazo',                 plazo_texto),
        ('Total invertido',       f'${total_invertido:,.2f}'),
    ]
    for idx, (lbl, val) in enumerate(params, 9):
        bg_ = P['ca'] if idx % 2 == 0 else P['c2']
        rh(ws1, idx, 20)
        cel(ws1, idx, 1, lbl,  bg_, P['gd'], sz=9, h='left',   brd=True)
        cel(ws1, idx, 2, val,  bg_, P['bl'], sz=9, bold=True,   brd=True)
        cel(ws1, idx, 3, '',   bg_, brd=False)
        cel(ws1, idx, 4, '',   bg_, brd=False)

    # ── Sección resultados ───────────────────────────────
    rh(ws1, 17, 8)
    merge_cel(ws1, 17, 1, 17, 4,
              val='📈  RESULTADOS DE LA SIMULACIÓN',
              bg=P['li'], fg=P['vl'], bold=True, sz=11, h='left')

    resultados_tbl = [
        ('Total invertido',         f'${total_invertido:,.2f}',  P['ce']),
        ('Valor final',             f'${valor_final:,.2f}',      P['ac']),
        ('Ganancia bruta',          f'${ganancia:,.2f}',         P['vl']),
        ('Retención en fuente 4%',  f'${retencion:,.2f}',        P['ro']),
        ('Saldo neto',              f'${saldo_neto:,.2f}',       P['ce']),
        ('Rentabilidad total',      f'{rentabilidad:.2f}%',      P['vl']),
        ('Ganancia sobre invertido',f'{pct_ganancia:.2f}%',      P['mp']),
    ]
    for idx, (lbl, val, color) in enumerate(resultados_tbl, 18):
        bg_ = P['ca'] if idx % 2 == 0 else P['c2']
        rh(ws1, idx, 20)
        cel(ws1, idx, 1, lbl, bg_, P['gd'], sz=9, h='left', brd=True)
        cel(ws1, idx, 2, val, bg_, color,   sz=9, bold=True, brd=True)
        cel(ws1, idx, 3, '',  bg_, brd=False)
        cel(ws1, idx, 4, '',  bg_, brd=False)

    # ── Barra de progreso ganancia ────────────────────────
    rh(ws1, 26, 8)
    merge_cel(ws1, 26, 1, 26, 8,
              val='📊  INDICADOR DE RENDIMIENTO',
              bg=P['li'], fg=P['gd'], bold=True, sz=11, h='left')

    pct_clip = min(100, pct_ganancia)
    color_barra = P['vl'] if pct_clip >= 20 else P['ac'] if pct_clip >= 5 else P['ro']
    rh(ws1, 27, 28)
    merge_cel(ws1, 27, 1, 27, 8,
              val=barra_pct(pct_clip),
              bg=P['c2'], fg=color_barra, sz=13, bold=True)
    ws1.cell(row=27, column=1).font = Font(
        color=color_barra, bold=True, size=13, name='Courier New'
    )

    rh(ws1, 28, 18)
    merge_cel(ws1, 28, 1, 28, 8,
              val=f'Por cada $1.000 invertido generaste ${round(ganancia/total_invertido*1000):,.0f} de ganancia bruta' if total_invertido > 0 else '',
              bg=P['ca'], fg=P['gd'], sz=9, italic=True)

    # ── Pie de hoja ──────────────────────────────────────
    rh(ws1, 30, 6)
    for i in range(1, 9):
        c = ws1.cell(row=30, column=i)
        c.fill = fl(P['mp'] if i <= 4 else P['cy'])

    rh(ws1, 31, 16)
    merge_cel(ws1, 31, 1, 31, 8,
              val='FinanBot  ·  Simulación educativa — Sin dinero real involucrado  ·  Colombia',
              bg=P['bg'], fg=P['mu'], sz=8, italic=True)


    # ════════════════════════════════════════════════════════
    #  HOJA 2 — PROYECCIÓN MES A MES
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('📅 Proyección')
    ws2.sheet_properties.tabColor = P['cy']
    fondo(ws2, 100, 10)

    # Cabecera
    rh(ws2, 1, 50)
    merge_cel(ws2, 1, 1, 1, 6,
              val='Proyección Mes a Mes — Interés Compuesto',
              bg=P['cy'], fg=P['bg'], bold=True, sz=16)

    rh(ws2, 2, 20)
    merge_cel(ws2, 2, 1, 2, 6,
              val=f'{nombre_tipo}  ·  Capital: ${capital:,.0f}  ·  Tasa: {tasa}%  ·  Plazo: {plazo_texto}',
              bg=P['ce'], fg=P['bg'], sz=9)

    rh(ws2, 3, 6)
    for i in range(1, 7):
        ws2.cell(row=3, column=i).fill = fl(P['cy'])

    # Encabezados tabla
    fila_enc(ws2, 4, [
        ('#',                   6),
        ('Período',             12),
        ('Valor c/ interés',    20),
        ('Capital invertido',   20),
        ('Interés acumulado',   20),
        ('Crecimiento %',       16),
    ], bg=P['li'], fg=P['gd'])

    # Datos mes a mes
    for p in proyeccion:
        r   = p['mes'] + 5
        bg_ = P['ca'] if p['mes'] % 2 == 0 else P['c2']
        pct_crec = round((p['valor'] / capital - 1) * 100, 2) if capital > 0 else 0
        col_crec = P['vl'] if pct_crec >= 0 else P['ro']
        rh(ws2, r, 19)

        cel(ws2, r, 1, p['mes'],       bg_, P['mu'],  sz=9)
        cel(ws2, r, 2, p['periodo'],   bg_, P['gr'],  sz=9)
        cel(ws2, r, 3, p['valor'],     bg_, P['ac'],  sz=9, bold=(p['mes'] == plazo))
        cel(ws2, r, 4, p['capital'],   bg_, P['ce'],  sz=9)
        cel(ws2, r, 5, p['intereses'], bg_, P['vl'],  sz=9)
        cel(ws2, r, 6, f'{pct_crec:.2f}%', bg_, col_crec, sz=9, bold=(p['mes'] == plazo))

        # Resaltar fila inicial y final
        if p['mes'] == 0 or p['mes'] == plazo:
            for col in range(1, 7):
                c = ws2.cell(row=r, column=col)
                c.fill   = fl(P['li'])
                c.font   = fn(P['bl'], bold=True, sz=9)
                c.border = bd(P['mp'])

    ws2.auto_filter.ref = 'A4:F4'

    # ── Gráfico líneas: valor vs capital ─────────────────
    # Datos auxiliares para el gráfico (columnas H, I, J)
    ws2['H1'] = 'Mes';    ws2['H1'].fill = fl(P['li']); ws2['H1'].font = fn(P['gd'], bold=True, sz=9)
    ws2['I1'] = 'Valor';  ws2['I1'].fill = fl(P['li']); ws2['I1'].font = fn(P['gd'], bold=True, sz=9)
    ws2['J1'] = 'Capital';ws2['J1'].fill = fl(P['li']); ws2['J1'].font = fn(P['gd'], bold=True, sz=9)
    cw(ws2, 8, 8); cw(ws2, 9, 16); cw(ws2, 10, 16)

    # Reducir puntos si hay muchos
    step = max(1, plazo // 24)
    puntos = [p for p in proyeccion if p['mes'] % step == 0 or p['mes'] == plazo]
    for i, p in enumerate(puntos, 2):
        ws2.cell(row=i, column=8, value=p['mes']).fill  = fl(P['c2'])
        ws2.cell(row=i, column=9, value=p['valor']).fill = fl(P['c2'])
        ws2.cell(row=i, column=10, value=p['capital']).fill = fl(P['c2'])
        for col in [8, 9, 10]:
            ws2.cell(row=i, column=col).font = fn(P['bl'], sz=8)

    max_row_chart = len(puntos) + 1

    lc = LineChart()
    lc.title  = 'Valor con interés vs Capital invertido'
    lc.style  = 10
    lc.width  = 26
    lc.height = 16

    data_line = Reference(ws2, min_col=9, max_col=10, min_row=1, max_row=max_row_chart)
    lc.add_data(data_line, titles_from_data=True)

    lc.series[0].graphicalProperties.line.solidFill = P['mp']
    lc.series[1].graphicalProperties.line.solidFill = P['ce']
    lc.series[0].graphicalProperties.line.width = 28000
    lc.series[1].graphicalProperties.line.width = 18000
    lc.series[1].graphicalProperties.line.dashDot = 'dash'

    ws2.add_chart(lc, 'H4')


    # ════════════════════════════════════════════════════════
    #  HOJA 3 — COMPARADOR DE ESCENARIOS
    # ════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('⚖️ Comparador')
    ws3.sheet_properties.tabColor = P['am']
    fondo(ws3, 50, 12)

    rh(ws3, 1, 50)
    merge_cel(ws3, 1, 1, 1, 7,
              val='Comparador de Escenarios de Inversión',
              bg=P['am'], fg=P['bg'], bold=True, sz=16)

    rh(ws3, 2, 20)
    merge_cel(ws3, 2, 1, 2, 7,
              val=f'Capital: ${capital:,.0f}  ·  Aporte mensual: ${aporte_mensual:,.0f}  ·  Plazo: {plazo_texto}',
              bg=P['ac'], fg=P['bg'], sz=9)

    rh(ws3, 3, 6)
    for i in range(1, 8):
        ws3.cell(row=3, column=i).fill = fl(P['am'])

    # Encabezados
    fila_enc(ws3, 4, [
        ('Tipo de inversión', 24),
        ('Tasa anual',        12),
        ('Valor final',       18),
        ('Ganancia',          16),
        ('Rentabilidad',      14),
        ('Vs. tu simulación', 18),
        ('Nivel de riesgo',   16),
    ], bg=P['li'], fg=P['gd'])

    colores_riesgo = {
        5:  (P['vl'], '🟢 Bajo'),
        8:  (P['vl'], '🟢 Bajo'),
        10: (P['ac'], '🟡 Medio'),
        15: (P['ro'], '🔴 Alto'),
        20: (P['re'], '🔴 Muy alto'),
    }

    for idx, e in enumerate(escenarios, 5):
        bg_   = P['li'] if e['es_actual'] else (P['ca'] if idx % 2 == 0 else P['c2'])
        fg_   = P['ac'] if e['es_actual'] else P['bl']
        diff  = round(e['valor_final'] - valor_final, 2)
        diff_s = f'+${diff:,.0f}' if diff >= 0 else f'-${abs(diff):,.0f}'
        col_d = P['vl'] if diff >= 0 else P['ro']
        col_r, txt_r = colores_riesgo.get(e['tasa'], (P['gr'], '⚪ Variable'))
        rh(ws3, idx, 22)

        cel(ws3, idx, 1, ('★ ' if e['es_actual'] else '') + e['nombre'],
            bg_, fg_, sz=9, bold=e['es_actual'], h='left')
        cel(ws3, idx, 2, f'{e["tasa"]}%',          bg_, P['ac'],  sz=9)
        cel(ws3, idx, 3, f'${e["valor_final"]:,.0f}', bg_, P['ce'], sz=9, bold=True)
        cel(ws3, idx, 4, f'${e["ganancia"]:,.0f}',    bg_, P['vl'], sz=9)
        cel(ws3, idx, 5, f'{e["rentabilidad"]:.2f}%', bg_, P['mp'], sz=9)
        cel(ws3, idx, 6, diff_s if not e['es_actual'] else '← Tu elección',
            bg_, col_d if not e['es_actual'] else P['ac'], sz=9, bold=True)
        cel(ws3, idx, 7, txt_r, bg_, col_r, sz=9)

    # ── Gráfico barras comparador ─────────────────────────
    ws3['I1'] = 'Escenario'; ws3['J1'] = 'Valor final'; ws3['K1'] = 'Ganancia'
    for c in ['I1','J1','K1']:
        ws3[c].fill = fl(P['li']); ws3[c].font = fn(P['gd'], bold=True, sz=9)
    cw(ws3, 9, 22); cw(ws3, 10, 16); cw(ws3, 11, 16)

    for i, e in enumerate(escenarios, 2):
        ws3.cell(row=i, column=9,  value=e['nombre']).fill       = fl(P['c2'])
        ws3.cell(row=i, column=10, value=e['valor_final']).fill  = fl(P['c2'])
        ws3.cell(row=i, column=11, value=e['ganancia']).fill     = fl(P['c2'])
        for col in [9, 10, 11]:
            ws3.cell(row=i, column=col).font = fn(P['bl'], sz=8)

    bc = BarChart()
    bc.type   = 'col'
    bc.title  = 'Valor final por escenario'
    bc.style  = 10
    bc.width  = 24
    bc.height = 16

    data_bar = Reference(ws3, min_col=10, max_col=11, min_row=1, max_row=len(escenarios)+1)
    cats_bar = Reference(ws3, min_col=9,  min_row=2,  max_row=len(escenarios)+1)
    bc.add_data(data_bar, titles_from_data=True)
    bc.set_categories(cats_bar)
    bc.series[0].graphicalProperties.solidFill = P['mp']
    bc.series[1].graphicalProperties.solidFill = P['vl']

    ws3.add_chart(bc, 'I4')


    # ════════════════════════════════════════════════════════
    #  HOJA 4 — ANÁLISIS FINANCIERO DETALLADO
    # ════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('💡 Análisis')
    ws4.sheet_properties.tabColor = P['vl']
    fondo(ws4, 50, 8)

    rh(ws4, 1, 50)
    merge_cel(ws4, 1, 1, 1, 6,
              val='Análisis Financiero Detallado',
              bg=P['ve'], fg=P['bl'], bold=True, sz=16)

    rh(ws4, 2, 20)
    merge_cel(ws4, 2, 1, 2, 6,
              val=f'Regla 50/30/20  ·  Fondo de emergencia  ·  Proyecciones de ahorro',
              bg=P['vl'], fg=P['bg'], sz=9)

    rh(ws4, 3, 6)
    for i in range(1, 7):
        ws4.cell(row=3, column=i).fill = fl(P['ve'])

    # ── Desglose de la ganancia ──────────────────────────
    rh(ws4, 5, 8)
    merge_cel(ws4, 5, 1, 5, 4,
              val='💰  DESGLOSE DE LA GANANCIA',
              bg=P['li'], fg=P['vl'], bold=True, sz=11, h='left')

    desglose = [
        ('Capital inicial aportado',    capital,         P['ce']),
        ('Aportes mensuales acumulados',aporte_mensual * plazo, P['gd']),
        ('Total invertido',             total_invertido, P['bl']),
        ('Intereses generados',         ganancia,        P['vl']),
        ('Retención en fuente (4%)',    -retencion,      P['ro']),
        ('Saldo neto final',            saldo_neto,      P['ac']),
    ]
    for idx, (lbl, val, color) in enumerate(desglose, 6):
        bg_ = P['ca'] if idx % 2 == 0 else P['c2']
        rh(ws4, idx, 20)
        cel(ws4, idx, 1, lbl,               bg_, P['gd'],  sz=9, h='left')
        cel(ws4, idx, 2, f'${val:,.2f}',    bg_, color,    sz=9, bold=True)
        pct_de_final = round(abs(val) / valor_final * 100, 1) if valor_final else 0
        cel(ws4, idx, 3, f'{pct_de_final}% del total', bg_, P['mu'], sz=8)
        cel(ws4, idx, 4, barra_pct(min(100, pct_de_final), 15),
            bg_, color, sz=8)
        ws4.cell(row=idx, column=4).font = Font(
            color=color, sz=8, name='Courier New'
        )
        cw(ws4, 1, 30); cw(ws4, 2, 18); cw(ws4, 3, 20); cw(ws4, 4, 24)

    # ── Proyecciones de ahorro mensual ───────────────────
    rh(ws4, 14, 8)
    merge_cel(ws4, 14, 1, 14, 4,
              val='📈  PROYECCIONES — Si ahorras mensualmente',
              bg=P['li'], fg=P['cy'], bold=True, sz=11, h='left')

    fila_enc(ws4, 15, [
        ('Horizonte',     16),
        ('Ahorro simple', 18),
        ('Con 8% EA',     18),
        ('Con tu tasa',   18),
    ], bg=P['li'], fg=P['gd'])

    def ci(t_anual, meses, aporte):
        tm = t_anual / 100 / 12
        b  = 0.0
        for _ in range(meses):
            b = b * (1 + tm) + aporte
        return round(b, 2)

    horizontes = [
        (6,   '6 meses'),
        (12,  '1 año'),
        (24,  '2 años'),
        (36,  '3 años'),
        (60,  '5 años'),
        (120, '10 años'),
    ]
    aporte_base = aporte_mensual if aporte_mensual > 0 else capital * 0.1

    for idx, (m, lbl) in enumerate(horizontes, 16):
        bg_    = P['ca'] if idx % 2 == 0 else P['c2']
        simple = round(aporte_base * m, 2)
        c8     = ci(8,   m, aporte_base)
        ct     = ci(tasa, m, aporte_base)
        mejor  = P['vl'] if ct > c8 else P['ac']
        rh(ws4, idx, 20)
        cel(ws4, idx, 1, lbl,             bg_, P['bl'],  sz=9, bold=True, h='left')
        cel(ws4, idx, 2, f'${simple:,.0f}', bg_, P['gr'],  sz=9)
        cel(ws4, idx, 3, f'${c8:,.0f}',    bg_, P['ce'],  sz=9)
        cel(ws4, idx, 4, f'${ct:,.0f}',    bg_, mejor,    sz=9, bold=True)

    # ── Consejo final ─────────────────────────────────────
    rh(ws4, 24, 8)
    merge_cel(ws4, 24, 1, 24, 4,
              val='💡  CONSEJO FINANCIERO',
              bg=P['li'], fg=P['ac'], bold=True, sz=11, h='left')

    consejo = (
        f'Con una tasa del {tasa}% anual durante {plazo_texto}, '
        f'tu dinero creció un {rentabilidad:.1f}%. '
        f'El interés compuesto generó ${ganancia:,.0f} adicionales sobre tu inversión. '
        f'Recuerda que el tiempo es tu mejor aliado: a mayor plazo, mayor rendimiento.'
    )
    rh(ws4, 25, 52)
    merge_cel(ws4, 25, 1, 25, 4,
              val=consejo,
              bg=P['ca'], fg=P['gd'], sz=9, italic=True, wrap=True)


    # ── Guardar y retornar ────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_str = datetime.now().strftime('%d-%m-%Y')
    nombre_archivo = (
        f'FinanBot_Simulacion_{nombre_tipo.replace(" ","_").replace("/","_")}_{fecha_str}.xlsx'
    )

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre_archivo
    )